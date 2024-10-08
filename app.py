from tkinter import *
from tkinter import Tk, Frame, Button, filedialog, messagebox
from openpyxl import load_workbook
from cpf import CPF

root = Tk()

class Application():
    def __init__(self):
        self.root = root
        self.cpfs = []

        self.screen()
        self.frames()
        self.buttons()
        root.mainloop()

    def screen(self):
        self.root.configure(background="#000")
        self.root.geometry("240x240")
        self.root.resizable(False, False)
        self.root.title("CPF Validator")

    def frames(self):
        self.container = Frame(self.root)
        self.container.place(relx=0.02, rely=0.02, relwidth=0.96, relheight=0.96) 

    def select_file(self):
        self.file_path = filedialog.askopenfilename(
            title="Selecione um arquivo Excel",
            filetypes=["Arquivos Excel", "*.xls *.xlsx"]
        )

        if self.file_path:
            self.open_file()
        else:
            messagebox.showinfo("Nenhum arquivo selecionado", "Nenhum arquivo foi selecionado")
            self.open_file()
    
    def buttons(self):
        self.button_open_file = Button(self.container, text="Selecione um arquivo Excel", command=self.select_file)
        self.button_open_file.place(relx=0.05, rely=0.4, relwidth=0.90, relheight=0.15)

    def open_file(self):
        self.wb = load_workbook(filename=self.file_path)
        self.wb_base = self.wb["base"]
        
        for row in self.wb_base.iter_rows(min_row=1, max_col=6, max_row=1, values_only=True):
            self.columns = row 
        self.column_cpf = self.columns.index("CPF")

        self.is_valid()

    def is_valid(self):
        for cell in self.wb_base.iter_cols(min_col=self.column_cpf + 1, max_col=self.column_cpf + 1, min_row=2, max_row=self.wb_base.max_row, values_only=True):
            for value in cell:
                 if value:
                    cpf_validator = CPF(str(value))
                    if cpf_validator.verify_cpf():
                        self.cpfs.append(True)
                    else:
                        self.cpfs.append(False)

    def insert_data_on_worksheet(self):
        self.one_column_after_cpf_column_index = self.self.collumn_cpf + 1 
        

Application()
    